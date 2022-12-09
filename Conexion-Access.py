import pyodbc
import pandas as pd
import xlwings as xw
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


# Conexión con Access
conexion = pyodbc.connect(
    r"DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};"
    r""r"DBQ=C:\Users\usuario\Documents\GitHub\Gestion_Pedidos_Ejercicios.accdb;"
)

consulta1 = "select * from clientes"
consulta2 = "select * from pedidos"
df_clientes = pd.read_sql(consulta1, conexion)
df_pedidos = pd.read_sql(consulta2, conexion)
conexion.close()


# Cruce de las tablas
df_merge = pd.merge(
    df_clientes,
    df_pedidos[['CÓDIGOCLIENTE', 'FECHADEPEDIDO']],
    on='CÓDIGOCLIENTE',
    how='left'
).fillna('Sin Información')


# Pegado de la tabla en Excel
wb = xw.Book(r'C:\Users\usuario\Documents\GitHub\Libro1.xlsx')
wb.app.screen_updating = False
wb.app.visible = False
wb.sheets["Hoja1"].clear_contents()
wb.sheets['Hoja1'].range(1, 1).value = df_merge.columns.tolist()
wb.sheets['Hoja1'].range(2, 1).value = df_merge.values
wb.save(r'C:\Users\usuario\Documents\GitHub\Libro1.xlsx')
wb.close()


# Formato tabla
libro = load_workbook(r'C:\Users\usuario\Documents\GitHub\Libro1.xlsx')
hoja = libro.get_sheet_by_name('Hoja1')
min_fila = hoja.min_row
max_fila = hoja.max_row
tab = Table(displayName="Table1", ref=f'A{min_fila}:H{max_fila}')
style = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=True
)

tab.tableStyleInfo = style
hoja.add_table(tab)
libro.save(r'C:\Users\usuario\Documents\GitHub\Libro1.xlsx')
